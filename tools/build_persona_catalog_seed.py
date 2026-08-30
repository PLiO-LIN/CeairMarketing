from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl


MODULE_KEYS = {
    "\u4e00\u3001\u57fa\u7840\u4fe1\u606f": "basic_info",
    "\u4e8c\u3001\u51fa\u884c\u884c\u4e3a": "travel_behavior",
    "\u4e09\u3001\u6d88\u8d39\u504f\u597d": "purchase_preferences",
    "\u56db\u3001\u4ef7\u503c\u4e0e\u5fe0\u8bda": "value_and_loyalty",
    "\u4e94\u3001\u5b9e\u65f6\u610f\u5411": "real_time_intent",
    "\u516d\u3001\u753b\u50cf\u5206\u7c7b\u7ed3\u679c": "persona_classification",
}

TYPE_KEYS = {
    "\u5b57\u7b26\u4e32": "string",
    "\u6587\u672c": "text",
    "\u679a\u4e3e": "enum",
    "\u591a\u9009\u679a\u4e3e": "multi_select",
    "\u65e5\u671f": "date",
    "\u65f6\u95f4\u6233": "timestamp",
    "\u6574\u6570": "integer",
    "\u6d6e\u70b9\u6570": "decimal",
    "\u5e03\u5c14": "boolean",
    "\u591a\u9009\u6587\u672c": "multi_text",
}

REQUIRED_MODES = {
    "\u662f": "required",
    "\u5426": "optional",
    "\u7cfb\u7edf\u81ea\u52a8": "system_generated",
}


def value(cell_value: object) -> object:
    return getattr(cell_value, "text", cell_value)


def clean(cell_value: object) -> str | None:
    cell_value = value(cell_value)
    if cell_value is None:
        return None
    text = str(cell_value).replace("\u00a0", " ").strip()
    return text or None


def persona_codes(raw: str | None) -> list[str]:
    if raw is None:
        return []
    if raw == "\u5168\u90e8":
        return ["ALL"]
    return [item.strip() for item in raw.split("/") if item.strip()]


def normalize_expression(raw: str | None) -> tuple[str | None, str | None, str | None]:
    if raw is None:
        return None, None, None
    expression = re.sub(r"\s+", " ", raw).strip()
    for symbol, operator in ((">=", "gte"), ("\u2265", "gte"), ("<=", "lte"), ("\u2264", "lte"), (">", "gt"), ("<", "lt")):
        if expression.startswith(symbol):
            return expression, operator, expression[len(symbol) :].strip()
    if expression.startswith("="):
        _, nested_operator, nested_value = normalize_expression(expression[1:].strip())
        if nested_operator is not None:
            return expression, nested_operator, nested_value
        return expression, "eq", expression[1:].strip()
    if expression.startswith("\u2208"):
        return expression, "in", expression[1:].strip()
    if "\u5305\u542b" in expression:
        return expression, "contains", expression.split("\u5305\u542b", 1)[1].strip()
    if re.match(r"^\d+\s*-\s*\d+", expression):
        return expression, "between", expression
    return expression, "expression", expression


def normalize_field_reference(field_code: str) -> tuple[str, str | None]:
    match = re.fullmatch(r"(.+)\uff08\u5386\u53f2\uff09", field_code)
    if match:
        return match.group(1), "historical"
    return field_code, None


def parse_dimensions(sheet: object) -> list[dict[str, object]]:
    current_module_name: str | None = None
    dimensions: list[dict[str, object]] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        cells = [clean(cell) for cell in row]
        module_name, field_name, field_code = cells[:3]
        if module_name and field_name is None and field_code is None:
            current_module_name = module_name
            continue
        if field_code is None:
            continue
        data_type = cells[3]
        required_text = cells[5]
        dimensions.append(
            {
                "module_key": MODULE_KEYS.get(current_module_name, f"module_{row_number}"),
                "module_name": current_module_name,
                "field_name": field_name,
                "field_code": field_code,
                "data_type": TYPE_KEYS.get(data_type, "unknown"),
                "source_data_type": data_type,
                "collection_method": cells[4],
                "required_mode": REQUIRED_MODES.get(required_text, "unknown"),
                "allowed_values": cells[6],
                "update_frequency": cells[7],
                "applicable_personas": persona_codes(cells[8]),
                "source_row": row_number,
            }
        )
    return dimensions


def parse_segments_and_rules(sheet: object, registered_codes: set[str]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    current_primary_code: str | None = None
    current_primary_name: str | None = None
    current_segment: dict[str, object] | None = None
    segments: list[dict[str, object]] = []
    rules: list[dict[str, object]] = []
    rule_order = 0

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [clean(cell) for cell in row]
        primary_name, subtype_name, segment_code = cells[:3]
        if row_number == 1:
            continue
        if primary_name:
            primary_parts = primary_name.split(".", 1)
            current_primary_code = primary_parts[0]
            current_primary_name = primary_parts[1] if len(primary_parts) == 2 else primary_name
        if segment_code:
            current_segment = {
                "segment_code": segment_code,
                "primary_persona_code": current_primary_code,
                "primary_persona_name": current_primary_name,
                "segment_name": subtype_name,
                "belongs_to": cells[3],
                "within_persona_share": cells[4],
                "recommended_products": cells[9],
                "recommended_channels": cells[10],
                "source_row": row_number,
            }
            segments.append(current_segment)
        field_code = cells[6]
        if current_segment is None or field_code is None:
            continue
        rule_order += 1
        expression, operator, condition_value = normalize_expression(cells[7])
        normalized_field_code, field_variant = normalize_field_reference(field_code)
        rules.append(
            {
                "segment_code": current_segment["segment_code"],
                "dimension_name": cells[5],
                "field_code": normalized_field_code,
                "field_variant": field_variant,
                "condition_expression": expression,
                "condition_operator": operator,
                "condition_value": condition_value,
                "data_source": cells[8],
                "field_registered": normalized_field_code in registered_codes,
                "rule_order": rule_order,
                "source_row": row_number,
            }
        )
    return segments, rules


def supplemental_dimensions() -> list[dict[str, object]]:
    return [
        {
            "module_key": "basic_info",
            "module_name": "\u4e00\u3001\u57fa\u7840\u4fe1\u606f",
            "field_name": "\u5e74\u9f84",
            "field_code": "age",
            "data_type": "integer",
            "source_data_type": "\u6d3e\u751f\u6574\u6570",
            "collection_method": "\u7cfb\u7edf\u8ba1\u7b97",
            "required_mode": "system_generated",
            "allowed_values": "\u7531\u51fa\u751f\u65e5\u671f\u8ba1\u7b97",
            "update_frequency": "\u5b9e\u65f6",
            "applicable_personas": ["ALL"],
            "source_row": 0,
            "is_supplemental": True,
        },
        {
            "module_key": "purchase_preferences",
            "module_name": "\u4e09\u3001\u6d88\u8d39\u504f\u597d",
            "field_name": "\u4ef7\u683c\u654f\u611f\u5ea6",
            "field_code": "price_sensitivity",
            "data_type": "enum",
            "source_data_type": "\u8865\u5145\u679a\u4e3e",
            "collection_method": "\u7cfb\u7edf\u5206\u6790",
            "required_mode": "system_generated",
            "allowed_values": "\u9ad8/\u4e2d/\u4f4e/\u62a2\u7968\u578b",
            "update_frequency": "\u6708\u5ea6",
            "applicable_personas": ["ALL"],
            "source_row": 0,
            "is_supplemental": True,
        },
    ]


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: build_persona_catalog_seed.py INPUT.xlsx OUTPUT.json")
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=False)
    dimensions_sheet = workbook["\u65c5\u5ba2\u753b\u50cf\u7ef4\u5ea6"]
    segments_sheet = workbook["\u753b\u50cf\u7c7b\u578b"]
    dimensions = parse_dimensions(dimensions_sheet)
    dimensions.extend(supplemental_dimensions())
    segments, rules = parse_segments_and_rules(segments_sheet, {item["field_code"] for item in dimensions})
    catalog = {
        "source_file": input_path.name,
        "catalog_version": "2026-08-30",
        "dimensions": dimensions,
        "segments": segments,
        "rules": rules,
    }
    output_path.write_text(json.dumps(catalog, ensure_ascii=True, indent=2), encoding="utf-8")
    print(json.dumps({"dimensions": len(dimensions), "segments": len(segments), "rules": len(rules), "unregistered_rule_fields": sorted({item["field_code"] for item in rules if not item["field_registered"]})}, ensure_ascii=True))


if __name__ == "__main__":
    main()
