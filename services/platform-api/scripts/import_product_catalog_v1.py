"""Idempotently import Eastern Airlines product packages from klg.xlsx."""
import argparse
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "services" / "platform-api"))
from app.database import Base, SessionLocal, engine
from app.db_models import ProductPackageRecord, TenantRecord

SOURCE_FILE = "klg.xlsx"
SOURCE_VERSION = "v1.0"


def clean(value):
    return "" if value is None else str(value).replace("/n", "\n").strip()


def parse_name(document_name):
    name = clean(document_name)
    if not name or name == "\u4ea7\u54c1\u4e00\u89c8" or "_" not in name:
        return None
    parts = [item.strip() for item in name.split("_") if item.strip()]
    if len(parts) >= 3 and parts[1].isdigit():
        return "_".join(parts[:2]), "_".join(parts[2:])
    return parts[0], "_".join(parts[1:])


def parse_products(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    grouped = OrderedDict()
    for row in workbook.active.iter_rows(min_row=2, values_only=True):
        parsed = parse_name(row[2])
        if not parsed:
            continue
        code, name = parsed
        grouped.setdefault(code, {"name": name, "parts": []})
        raw = clean(row[0])
        if raw and raw not in grouped[code]["parts"]:
            grouped[code]["parts"].append(raw)
    products = []
    for code, item in grouped.items():
        merged = "\n\n".join(item["parts"])
        description_match = re.search(r"\u4ea7\u54c1\u63cf\u8ff0\uff1a(.+?)(?:\u4ea7\u54c1\u63d0\u793a\uff1a|$)", merged, re.S)
        hint_match = re.search(r"\u4ea7\u54c1\u63d0\u793a\uff1a(.+)", merged, re.S)
        products.append({"external_id": "CEA-" + code, "name": item["name"], "description": (description_match.group(1).strip() if description_match else merged[:1800]), "eligibility": (hint_match.group(1).strip() if hint_match else merged[:1000])})
    return products


def import_products(path, tenant_code, upsert=False):
    products = parse_products(path)
    Base.metadata.create_all(bind=engine)
    result = {"created": 0, "updated": 0, "skipped": 0, "total": len(products)}
    with SessionLocal() as db:
        tenant = db.scalar(select(TenantRecord).where(TenantRecord.code == tenant_code))
        if tenant is None:
            raise ValueError("\u79df\u6237\u4e0d\u5b58\u5728\uff1a" + tenant_code)
        for item in products:
            record = db.scalar(select(ProductPackageRecord).where(ProductPackageRecord.tenant_id == tenant.id, ProductPackageRecord.external_id == item["external_id"]))
            payload = {"name": item["name"], "product_type": "\u822a\u7a7a\u8425\u9500\u4ea7\u54c1\u5305", "description": item["description"][:2000], "eligibility": item["eligibility"][:1000], "version": SOURCE_VERSION, "status": "\u8349\u7a3f"}
            if record is None:
                db.add(ProductPackageRecord(tenant_id=tenant.id, external_id=item["external_id"], created_by=None, **payload))
                result["created"] += 1
            elif upsert:
                for key, value in payload.items():
                    setattr(record, key, value)
                result["updated"] += 1
            else:
                result["skipped"] += 1
        db.commit()
    return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="\u5bfc\u5165\u4e1c\u822a\u4ea7\u54c1\u77e5\u8bc6\u76ee\u5f55")
    parser.add_argument("--file", type=Path, default=ROOT / "docs" / SOURCE_FILE)
    parser.add_argument("--tenant", default="CEA-HQ")
    parser.add_argument("--upsert", action="store_true")
    args = parser.parse_args()
    print(json.dumps(import_products(args.file, args.tenant, args.upsert), ensure_ascii=False, indent=2))
