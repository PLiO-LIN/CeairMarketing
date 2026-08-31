"""Import the Eastern Airlines persona catalog idempotently."""
import argparse
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "services" / "platform-api"))
from app.database import Base, SessionLocal, engine
from app.db_models import AudienceTagRecord, PersonaDimensionDefinitionRecord, PersonaSegmentRecord, PersonaSegmentRuleRecord, TenantRecord

SRC = "画像维度及类型(1).xlsx"
VER = "v1.0"
DHEAD = ["采集模块", "字段名称", "字段代码", "数据类型", "采集方式", "是否必填", "选项/格式", "更新频率", "适用画像"]
SHEAD = ["主画像", "子类型", "代码", "画像归属", "占比（主画像内）", "判定维度", "采集字段", "判定阈值/条件", "数据来源", "最匹配营销产品", "建议触达渠道"]

def clean(value):
    return "" if value is None else str(value).strip()

def parse(path):
    wb = load_workbook(path, data_only=False, read_only=True)
    drows = list(wb["旅客画像维度"].iter_rows(values_only=True))
    if [clean(x) for x in drows[0]] != DHEAD:
        raise ValueError("旅客画像维度表头不一致")
    dims, module = [], ""
    for row_no, row in enumerate(drows[1:], 2):
        v = [clean(x) for x in row]
        if v[0] and not v[2]:
            module = v[0]
        elif v[2]:
            dims.append(dict(module_key=module.split("、", 1)[0].lower() or "profile", module_name=module, field_name=v[1], field_code=v[2], data_type=v[3] or "string", source_data_type=v[3], collection_method=v[4], required_mode={"是": "required", "否": "optional", "系统自动": "system"}.get(v[5], v[5] or "optional"), allowed_values=v[6], update_frequency=v[7], applicable_personas=[x.strip() for x in v[8].replace("，", ",").split(",") if x.strip()] or ["ALL"], source_row=row_no + 1))
    srows = list(wb["画像类型"].iter_rows(values_only=True))
    if [clean(x) for x in srows[0]] != SHEAD:
        raise ValueError("画像类型表头不一致")
    segs, rules, primary, primary_code, segment_code = [], [], "", "", ""
    for row_no, row in enumerate(srows[1:], 2):
        v = [clean(x) for x in row]
        if v[0]:
            primary, primary_code = v[0], v[0].split(".", 1)[0]
        if v[1] and v[2]:
            segment_code = v[2]
            segs.append(dict(segment_code=segment_code, primary_persona_code=primary_code, primary_persona_name=primary.split(".", 1)[-1], segment_name=v[1], belongs_to=v[3], within_persona_share=float(v[4]) if v[4] else None, recommended_products=v[9], recommended_channels=v[10], source_row=row_no + 1))
        if v[5] and segment_code:
            rules.append(dict(segment_code=segment_code, dimension_name=v[5], field_code=v[6], field_variant="", condition_expression=v[7], condition_operator="expression", condition_value=v[7], data_source=v[8], field_registered=bool(v[6]), rule_order=sum(x["segment_code"] == segment_code for x in rules) + 1, source_row=row_no + 1))
    return dims, segs, rules

def import_catalog(path, tenant_code, upsert=False):
    dims, segs, rules = parse(path)
    Base.metadata.create_all(bind=engine)
    result = {key: 0 for key in ["dimensions_created", "dimensions_updated", "segments_created", "segments_updated", "rules_created", "tags_created", "tags_updated"]}
    with SessionLocal() as db:
        tenant = db.scalar(select(TenantRecord).where(TenantRecord.code == tenant_code))
        if tenant is None:
            raise ValueError("租户不存在：" + tenant_code)
        segment_records = {}
        for item in dims:
            record = db.scalar(select(PersonaDimensionDefinitionRecord).where(PersonaDimensionDefinitionRecord.tenant_id == tenant.id, PersonaDimensionDefinitionRecord.field_code == item["field_code"]))
            payload = {key: item[key] for key in ["module_key", "module_name", "field_name", "field_code", "data_type", "source_data_type", "collection_method", "required_mode", "allowed_values", "update_frequency", "source_row"]}
            payload.update(tenant_id=tenant.id, source_file=SRC, source_version=VER, applicable_personas_json=json.dumps(item["applicable_personas"], ensure_ascii=False), is_supplemental=False)
            if record is None:
                db.add(PersonaDimensionDefinitionRecord(**payload)); result["dimensions_created"] += 1
            elif upsert:
                for key, value in payload.items():
                    if key != "tenant_id": setattr(record, key, value)
                result["dimensions_updated"] += 1
            desc = "；".join(filter(None, ["数据类型：" + item["source_data_type"], "采集方式：" + item["collection_method"], "必填：" + item["required_mode"], "选项/格式：" + item["allowed_values"], "更新频率：" + item["update_frequency"], "适用画像：" + "、".join(item["applicable_personas"]), "来源文件：" + SRC]))[:1000]
            tag = db.scalar(select(AudienceTagRecord).where(AudienceTagRecord.tenant_id == tenant.id, AudienceTagRecord.code == item["field_code"]))
            tag_data = dict(name=item["field_name"], category=item["module_name"] or "旅客画像维度", source="画像维度目录导入", description=desc, enabled=True)
            if tag is None:
                db.add(AudienceTagRecord(tenant_id=tenant.id, code=item["field_code"], **tag_data)); result["tags_created"] += 1
            elif upsert:
                for key, value in tag_data.items(): setattr(tag, key, value)
                result["tags_updated"] += 1
        db.flush()
        for item in segs:
            record = db.scalar(select(PersonaSegmentRecord).where(PersonaSegmentRecord.tenant_id == tenant.id, PersonaSegmentRecord.segment_code == item["segment_code"]))
            payload = dict(item, tenant_id=tenant.id, source_file=SRC, source_version=VER)
            if record is None:
                record = PersonaSegmentRecord(**payload); db.add(record); db.flush(); result["segments_created"] += 1
            elif upsert:
                for key, value in payload.items():
                    if key != "tenant_id": setattr(record, key, value)
                result["segments_updated"] += 1
            segment_records[item["segment_code"]] = record
            code = "persona_" + item["segment_code"]
            tag = db.scalar(select(AudienceTagRecord).where(AudienceTagRecord.tenant_id == tenant.id, AudienceTagRecord.code == code))
            tag_data = dict(name=item["segment_name"], category="画像类型/" + item["primary_persona_name"], source="画像类型目录导入", description=("主画像：" + item["primary_persona_name"] + "；归属：" + item["belongs_to"] + "；匹配产品：" + item["recommended_products"] + "；触达渠道：" + item["recommended_channels"] + "；来源文件：" + SRC)[:1000], enabled=True)
            if tag is None:
                db.add(AudienceTagRecord(tenant_id=tenant.id, code=code, **tag_data)); result["tags_created"] += 1
            elif upsert:
                for key, value in tag_data.items(): setattr(tag, key, value)
                result["tags_updated"] += 1
        db.flush()
        for item in rules:
            segment = segment_records.get(item["segment_code"])
            if segment and db.scalar(select(PersonaSegmentRuleRecord).where(PersonaSegmentRuleRecord.segment_id == segment.id, PersonaSegmentRuleRecord.source_row == item["source_row"])) is None:
                db.add(PersonaSegmentRuleRecord(segment_id=segment.id, **{key: value for key, value in item.items() if key != "segment_code"})); result["rules_created"] += 1
        db.commit()
    return result

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="导入东航旅客画像目录")
    parser.add_argument("--file", type=Path, default=ROOT / "docs" / SRC)
    parser.add_argument("--tenant", default="CEA-HQ")
    parser.add_argument("--upsert", action="store_true")
    args = parser.parse_args()
    print(json.dumps(import_catalog(args.file, args.tenant, args.upsert), ensure_ascii=False, indent=2))
